"""Выгрузка маршрутного листа доставки в Excel."""

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

HEADERS = ["Контакт", "Телефон", "Товар", "Цена", "Количество", "ЖК"]


def _thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def build_manifest_rows(orders: list[Any]) -> list[dict]:
    rows = []
    for order in orders:
        contact = order.user.full_name if order.user else "—"
        phone = order.user.phone if order.user else ""
        for item in order.items:
            rows.append({
                "contact": contact,
                "phone": phone,
                "product": item.product_name,
                "price": item.price,
                "quantity": item.quantity,
                "address": order.address,
            })
    rows.sort(key=lambda r: (
        r["contact"].lower(),
        r["address"].lower(),
        r["product"].lower(),
    ))
    return rows


def manifest_to_xlsx(rows: list[dict], sheet_title: str = "Доставки") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = _thin_border()

    for i, row in enumerate(rows, 2):
        values = [
            row["contact"],
            row["phone"],
            row["product"],
            row["price"],
            row["quantity"],
            row["address"],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = _thin_border()
            if col == 4:
                cell.number_format = "#,##0"

    widths = [28, 18, 36, 12, 14, 28]
    for i, w in enumerate(widths, 1):
        col_letter = chr(64 + i) if i <= 26 else "A"
        ws.column_dimensions[col_letter].width = w

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
